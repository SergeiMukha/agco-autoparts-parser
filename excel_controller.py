from openpyxl import load_workbook

class ExcelController:
    def __init__(self, filename, row):
        self.row = row
        self.filename = filename

        self.workbook = load_workbook(self.filename)
        self.sheet = self.workbook.active
        self.max_row = self.sheet.max_row
    
    def save_document(self):
        self.workbook.save(self.filename)